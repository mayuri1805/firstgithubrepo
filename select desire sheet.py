import openpyxl
path = "C:\\Users\Asus\\OneDrive\\Documents\\imagelogo.xl.xlsx"
# Creat workbook (Excelsheet Object)load_
workbook = openpyxl.load_workbook(path)
sheet = workbook.create_sheet(" Sheet 1")

# sheet = workbook["Select sheet"]

rows= sheet.max_row
cols= sheet.max_column
roll= ["Manager1","Admin1","Staff1","Student1"]
user= ["xmngr1234","xadmin1234","xstaff@4567","xstud@987"]
password= ["a34555","b5656565","c787878","d977675"]

for r in  range(1,len(roll)+1):
    sheet.cell(row=rows+r,column=1).value=(rows-1)+r
    sheet.cell(row=rows+r,column=2).value=roll[rows-1]
    sheet.cell(row=rows+r,column=3).value=user[rows-1]
    sheet.cell(row=rows+r,column=4).value=password[rows-1]
workbook.save(path)



import openpyxl
wb=openpyxl.load_workbook('D:\\xl\\xlpracticle.xlsx')
sheet=wb.active
r=sheet.max_row
c=sheet.max_column
print(r)
print(c)
for i in range(r+1):
    for j in range(c+1):
        print(sheet.cell(row=i+1,column=j+1).value,end=' ')
    print()
for i in range(r+1,r+6):
    for j in range(c+1):
        sheet.cell(row=i,column=j+1).value="hello"
wb.save("D:\\xl\\xlpracticle.xlsx")

for i in range(r+21):
    for j in range(c+1):
        sheet.cell(row=i,column=j+1).value="excel"
wb.save("D:\\xl\\xlpracticle.xlsx")
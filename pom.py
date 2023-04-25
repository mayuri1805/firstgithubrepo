import openpyxl

username=''
password=''


path = "C:\\Users\Asus\\OneDrive\\Documents\\pom.xl.xlsx"

# Creat workbook (Excelsheet Object)load_
workbook = openpyxl.load_workbook(path)

# Use following function when excel sheet contain more than one workbook
# sheet = workbook.get_sheet_by_name("sheet1")

# For single worksheet
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)
:
    if
for i in range(1,rows) sheet.cell(row=i+1,column=2).value == "manager":
        username=sheet.cell(row=i+1,column=3).value
        password=sheet.cell(row=i+1,column=4).value

# import openpyxl
# # copy xl sheet path
# path = "C:\\Users\\Asus\\OneDrive\\Desktop\\data phase.xlsx"
# # Creat workbook (Excelsheet Object)load_
# workbook = openpyxl.load_workbook(path)

# for single worksheet
# sheet= workbook.active
# rows= sheet.max_row
# cols= sheet.max_column
# roll= ["Manager1","Admin1","Staff1","Student1"]
# user= ["xmngr1234","xadmin1234","xstaff@4567","xstud@987"]
# password= ["a34555","b5656565","c787878","d977675"]
#
# for r in  range(1,len(roll)+1):
#     sheet.cell(row=rows+r,column=1).value=(rows-1)+r
#     sheet.cell(row=rows+r,column=2).value=roll[rows-1]
#     sheet.cell(row=rows+r,column=3).value=user[rows-1]
#     sheet.cell(row=rows+r,column=4).value=password[rows-1]
# workbook.save(path)


# Reading data from xlsheet

import openpyxl
path = "C:\\Users\Asus\\OneDrive\\Desktop\\data phase.xlsx"
# Creat workbook (Excelsheet Object)load_
workbook = openpyxl.load_workbook(path)

# Use following function when excel sheet contain more than one workbook
# sheet = workbook.get_sheet_by_name("sheet1")

# For single worksheet
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)

for i in range(1,rows):
    print(sheet.cell(rows=i+1,column=1).value,end=" _ ")
    print(sheet.cell(rows=i+1,column=2).value,end=" _ ")
    print(sheet.cell(rows=i+1,column=3).value,end=" _ ")
    print(sheet.cell(rows=i+1,column=4).value,end=" _ ")
    print()

# Read the data dynamically with iterative column
# for i in range(i,rows):
#     for j in range(1,cols):
#         print(sheet.cell(row=i+1,column=j).value,end=" - " )
#     print()
#








